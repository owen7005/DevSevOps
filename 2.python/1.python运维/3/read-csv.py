#!/usr/bin/env python-devops
# -*- coding: utf-8 -*-
# @File  : read-csv.py
# @Author: Anthony.waa
# @Date  : 2019/3/15 0015
# @Desc  : PyCharm

import openpyxl

def readExel():
    filename = r'C:\Users\Administrator\Desktop\guest\files\userCase.xlsx'
    inwb = openpyxl.load_workbook(filename)  # 读文件

    sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    print(ws.cell(1,1).value)



    for r in range(1, rows):
        for c in range(1, cols):
            print(ws.cell(r, c).value)
        # if r == 10:
        #     break

readExel()

# import os
# print(os.path.realpath(os.getcwd()))
# print(os.getcwd())