'''
openpyxl模块：第三方模块
    - 可以对Excle表格进行操作的模块

    - 下载:
        pip3 install openpyxl

    - Excel版本:
        2003之前:
            excle名字.xls

        2003以后:
            excle名字.xlsx

    - 清华源: https://pypi.tuna.tsinghua.edu.cn/simple

    - 配置永久第三方源:
        D:\Python36\Lib\site-packages\pip\_internal\models\index.py

'''

# 写入数据
# from openpyxl import Workbook
# # 获取Excel文件对象
# wb_obj = Workbook()
#
# wb1 = wb_obj.create_sheet('python13期工作表1', 1)
# wb2 = wb_obj.create_sheet('python13期工作表2', 2)
#
# # 修改工作表名字: 为python13期工作表1标题修改名字 ---》 tank大宝贝
# print(wb1.title)
# wb1.title = 'tank大宝贝'
# print(wb1.title)

# 为第一张工作表添加值
# wb1['工作簿中的表格位置']
# wb1['A10'] = 200
# wb1['B10'] = 1000
# wb1['C10'] = '=SUM(A10:B10)'
#
#
# wb2['A1'] = 100
#
# # 生成Excel表格
# wb_obj.save('python13期.xlsx')
# print('excel表格生成成功')


# 读取数据
# from openpyxl import load_workbook
# wb_obj = load_workbook('python13期.xlsx')
# print(wb_obj)
#
# # wb_obj['表名']
# wb1 = wb_obj['tank大宝贝']
# print(wb1['A10'].value)
# wb1['A10'] = 20
# print(wb1['A10'].value)




# 批量写入100条数据
from openpyxl import Workbook

wb_obj = Workbook()

wb1 = wb_obj.create_sheet('工作表1')

# wb1['表格位置'] = 对应的值
# n = 1
# for line in range(100):
#
#     wb1['A%s' % n] = line + 1
#     n += 1

# 假设: 一万条数据的字典
dict1 = {
    'name': 'tank',
    'age': 17
}

n = 1
init_val = 65
chr(65)  # 65-90
for key, value in dict1.items():
    wb1['A%s' % n] = key
    wb1['B%s' % n] = value
    n += 1

wb_obj.save('批量插入的数据2.xlsx')



